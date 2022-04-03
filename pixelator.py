import time
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Color
from PIL import Image
from PIL import ImageColor
from os.path import exists

excel_color = {
    "DarkRed": "be0039",
    "Red": "ff4500",
    "Orange": "ffa800",
    "Yellow": "ffd635",
    "DarkGreen": "00a368",
    "Green": "00cc78",
    "LightGreen": "00cc78",
    "DarkTeal": "00756f",
    "Teal": "009eaa",
    "DarkBlue": "2450a4",
    "Blue": "3690ea",
    "LightBlue": "51e9f4",
    "Indigo": "493ac1",
    "Periwinkle": "6a5cff",
    "DarkPurple": "811e9f",
    "Purple": "b44ac0",
    "Pink": "ff3881",
    "LightPink": "ff99aa",
    "DarkBrown": "6d482f",
    "Brown": "9c6926",
    "Black": "000000",
    "DarkGrey": "898d90",
    "LightGrey": "d4d7d9",
    "White": "ffffff"
}

main_color = {key.upper(): ImageColor.getrgb(f"#{value}ff") for key, value in excel_color.items()}
main_color["BLANK"] = ImageColor.getrgb("#00000000")
inverse_color = {value: key for key, value in main_color.items()}
excel_color = {key.upper(): f"ff{value}" for key, value in excel_color.items()}
excel_color["BLANK"] = "00000000"

all_colors = list(main_color.values())
weights = {'R': 1.0, 'G': 1.0, 'B': 1.0, 'A': 1.0}

def user_change_weights():
    global weights
    for key in weights.keys():
        weights[key] = get_user_float(f"New weight for channel \'{key}\': ")

def get_user_img_size(bounds_x: tuple = (0, float("inf")), bounds_y: tuple = (0, float("inf"))) -> tuple:
    tgt_dims = ("width", "height")
    tgt_size = [-1, -1]
    bounds = bounds_x
    for i in range(len(tgt_size)):
        # while out of bounds
        while tgt_size[i] < bounds[0] or tgt_size[i] > bounds[1]:
            try:
                req_s = int(input(f"Input desired {tgt_dims[i]}: "))
            except ValueError as ve:
                print(f"Invalid input: {ve}")
                continue
            tgt_size[i] = req_s
            bounds = bounds_y
    return tuple(tgt_size)

def get_user_int(prompt: str) -> int:
    while(True):
        try:
            user_in = int(input(prompt))
        except ValueError as ve:
            print(f"Invalid input: {ve}")
            continue
        return user_in

def get_user_float(prompt: str) -> float:
    while(True):
        try:
            user_in = float(input(prompt))
        except ValueError as ve:
            print(f"Invalid input: {ve}")
            continue
        return user_in

def get_user_bool(prompt: str) -> bool:
    return input(prompt).upper() == 'Y'

def get_user_file(prompt: str) -> str:
    user_in = input(prompt)
    return get_file(user_in)

def get_file(file_loc: str) -> str:
    if exists(file_loc):
        return file_loc
    
def color_closest(source: list) -> list:
    min_dist = float('INF')
    best_color = source
    global all_colors
    global weights
    weightss = list(weights.values())
    try:
        if source[3] == 0:
            return source[3]
    except IndexError:
        pass
    for colorr in all_colors:
        distance = 0
        for i in range(len(source)):
            distance += ((colorr[i]-source[i])*(1/weightss[i]))**2
        distance = distance**0.5
        if distance < min_dist:
            min_dist = distance
            best_color = colorr[:len(source)]
    return best_color

class rplace_image:
    def __init__(self, image: str):
        '''Initialize with file location'''
        try:
            IMG = Image.open(image)
        except:
            return None
        self.img = IMG
        self.x = IMG.size[0]
        self.y = IMG.size[1]
        self.outs = dict()

    def pixelate(self, x_size: int, y_size: int):
        src_w, src_h = self.img.size
        x_size = round(x_size)
        y_size = round(y_size)
        w_offset = round(src_w / x_size)
        h_offset = round(src_h / y_size)
        src_img = np.asarray(self.img)
        tgt_img = np.empty(shape=(y_size, x_size, len(src_img[0][0])), dtype=src_img.dtype)
        for x_tgt in range(0, x_size):
            for y_tgt in range(0, y_size):
                avg_color = [0]*len(self.img.mode)
                count = 0
                # sample the area
                for x_src in range(x_tgt*w_offset, (x_tgt*w_offset)+w_offset):
                    for y_src in range(y_tgt*h_offset, (y_tgt*h_offset)+h_offset):
                        error = False
                        for i in range(len(avg_color)):
                            try:
                                avg_color[i] += src_img[y_src][x_src][i]
                            except IndexError as ie:
                                error = True
                                break
                        if not error:
                            count += 1
                avg_color = color_closest([item//count for item in avg_color]) # average rgb values
                tgt_img[y_tgt][x_tgt] = avg_color
        tgt_out = Image.fromarray(tgt_img)
        self.outs[f"{x_size}x{y_size}"] = tgt_out

    def pixelate_square(self, square: int):
        return self.pixelate(self.x / square, self.y / square)

def main():
    file = get_user_file("Specify image: ")
    if not file:
        return None
    px = rplace_image(file)
    #size = get_user_img_size((0, px.x-1), (0, px.y-1))
    #out = px.pixelate(size[0], size[1])
    # get color restrictions
    if get_user_bool("Specify color restrictions? Y/N: "):
        print("Possible colors: ")
        print(list(main_color.keys()), sep=", ")
        print("Specify the colors to allow by name sperated by commas: ")
        color_filter_user = input()
        color_filters = [f"{value.lstrip().strip()}" for value in color_filter_user.upper().split(',')]
        if not color_filters:
            print("No filters applied.")
        global all_colors
        all_colors = list()
        for c_filter in color_filters:
            try:
                all_colors.append(main_color[c_filter])
            except KeyError as ke:
                print(f"Invalid color \"{c_filter}\": {ke}")
                continue
    if get_user_bool("Specify color channel weights? Y/N: "):
        print("Note: Alpha channel not applicable to all images")
        print("Input RGBA color weights: ")
        user_change_weights()
    # get sizes
    sizes = input("Square sizes (int), separated by comma: ")
    sizes = sizes.strip().split(',')
    for index, size in enumerate(sizes):
        try:
            sizes[index] = int(sizes[index].lstrip().strip())
        except ValueError as ve:
            print(f"Invalid size in your input: {ve}")
            continue
    sizes = [value for value in sizes if type(value) == int]
    start = time.time()
    for size in sizes:
        px.pixelate_square(size)
    end = time.time()
    print(f"Rendered {len(px.outs)} images in {(end-start)*1000:.1f} ms")
    out_path = input("Name to save to (do not include extension): ")
    if excel := get_user_bool(f"Output to Excel? Y/N: "):
        count = 0
        wb = Workbook()
        wb.remove(wb.active)
        for key, img in px.outs.items():
            img_ar = np.asarray(img)
            ws = wb.create_sheet(f"{count}-{key}")
            for row in range(len(img_ar)):
                for column in range(len(img_ar[0])):
                    colour = Color(excel_color[inverse_color[tuple(img_ar[row][column])]])
                    fill = PatternFill(fill_type='solid', fgColor=colour)
                    ws.cell(row=row+1, column=column+1).fill = fill
            count += 1
    for key, out in px.outs.items():
        out.save(f"img/{out_path}-{key}.png", format='PNG')
    if excel:
        wb.save(f"img/{out_path}.xlsx")

if __name__ == "__main__":
    main()